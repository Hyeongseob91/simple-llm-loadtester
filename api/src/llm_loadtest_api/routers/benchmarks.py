"""Benchmark API routes."""

import csv
import io
from datetime import datetime
from typing import Literal, Optional

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import Response

from llm_loadtest_api.auth import APIKeyAuth
from llm_loadtest_api.database import Database
from llm_loadtest_api.logging_config import get_logger, log_benchmark_event
from llm_loadtest_api.services.benchmark_service import BenchmarkService
from llm_loadtest_api.models.schemas import (
    BenchmarkRequest,
    BenchmarkResponse,
    BenchmarkStatus,
    CompareRequest,
    CompareResponse,
    HealthResponse,
    RunListResponse,
)
from llm_loadtest_api import __version__

# Logger
logger = get_logger(__name__)


router = APIRouter(prefix="/api/v1/benchmark", tags=["benchmark"])

# Database and service instances (singleton pattern)
_db: Optional[Database] = None
_service: Optional[BenchmarkService] = None


def get_db() -> Database:
    """Get database instance."""
    global _db
    if _db is None:
        _db = Database()
    return _db


def get_service() -> BenchmarkService:
    """Get benchmark service instance."""
    global _service
    if _service is None:
        _service = BenchmarkService(get_db())
    return _service


@router.get("/health", response_model=HealthResponse)
async def health_check() -> HealthResponse:
    """Check API health."""
    return HealthResponse(
        status="healthy",
        version=__version__,
        timestamp=datetime.now(),
    )


@router.post("/run", response_model=dict, dependencies=[Depends(APIKeyAuth(required=True))])
async def start_benchmark(
    request: BenchmarkRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Start a new benchmark run.

    Requires API key authentication via X-API-Key header.

    Returns the run_id for tracking progress.
    """
    logger.info(
        "benchmark_request_received",
        server_url=request.server_url,
        model=request.model,
        adapter=request.adapter,
    )

    run_id = await service.start_benchmark(request)

    log_benchmark_event(
        "benchmark_started",
        run_id=run_id,
        extra={
            "server_url": request.server_url,
            "model": request.model,
            "adapter": request.adapter,
        },
    )

    return {"run_id": run_id, "status": "started"}


@router.get("/run/{run_id}", response_model=BenchmarkStatus)
async def get_run_status(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> BenchmarkStatus:
    """Get benchmark run status."""
    run = service.get_status(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    return BenchmarkStatus(
        run_id=run["id"],
        status=run["status"],
        server_url=run["server_url"],
        model=run["model"],
        adapter=run["adapter"],
        started_at=run.get("started_at"),
        completed_at=run.get("completed_at"),
        created_at=run["created_at"],
    )


@router.get("/result/{run_id}")
async def get_run_result(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Get benchmark result."""
    result = service.get_result(run_id)
    if not result:
        # Check if run exists but not completed
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    return result


@router.get("/history", response_model=RunListResponse)
async def list_runs(
    limit: int = 50,
    offset: int = 0,
    status: Optional[str] = None,
    service: BenchmarkService = Depends(get_service),
) -> RunListResponse:
    """List benchmark runs."""
    runs = service.list_runs(limit, offset, status)

    return RunListResponse(
        runs=[
            BenchmarkStatus(
                run_id=r["id"],
                status=r["status"],
                server_url=r["server_url"],
                model=r["model"],
                adapter=r["adapter"],
                started_at=r.get("started_at"),
                completed_at=r.get("completed_at"),
                created_at=r["created_at"],
            )
            for r in runs
        ],
        total=len(runs),  # TODO: Get actual total count
        limit=limit,
        offset=offset,
    )


@router.delete("/run/{run_id}", dependencies=[Depends(APIKeyAuth(required=True))])
async def delete_run(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Delete a benchmark run.

    Requires API key authentication via X-API-Key header.
    """
    logger.info("benchmark_delete_request", run_id=run_id)

    deleted = service.delete_run(run_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Run not found")

    log_benchmark_event("benchmark_deleted", run_id=run_id)

    return {"deleted": run_id}


@router.post("/compare")
async def compare_runs(
    request: CompareRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Compare multiple benchmark runs."""
    comparison = service.compare_runs(request.run_ids)
    return comparison


@router.get("/result/{run_id}/export")
async def export_result(
    run_id: str,
    format: Literal["csv", "xlsx"] = "csv",
    service: BenchmarkService = Depends(get_service),
) -> Response:
    """Export benchmark result to CSV or Excel format.

    Args:
        run_id: The benchmark run ID.
        format: Export format ('csv' or 'xlsx').

    Returns:
        File download response.
    """
    result = service.get_result(run_id)
    if not result:
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    if format == "csv":
        content = _export_to_csv(result)
        return Response(
            content=content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.csv"'
            },
        )
    else:
        content = _export_to_xlsx(result)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.xlsx"'
            },
        )


def _export_to_csv(result: dict) -> str:
    """Convert benchmark result to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata section
    writer.writerow(["# Benchmark Result Export"])
    writer.writerow(["Run ID", result.get("run_id", "N/A")])
    writer.writerow(["Model", result.get("model", "N/A")])
    writer.writerow(["Server URL", result.get("server_url", "N/A")])
    writer.writerow(["Adapter", result.get("adapter", "N/A")])
    writer.writerow(["Started At", result.get("started_at", "N/A")])
    writer.writerow(["Completed At", result.get("completed_at", "N/A")])
    writer.writerow(["Duration (s)", result.get("duration_seconds", "N/A")])
    writer.writerow([])

    # Summary section
    summary = result.get("summary", {})
    writer.writerow(["# Summary"])
    writer.writerow(["Best Throughput (tok/s)", summary.get("best_throughput", "N/A")])
    writer.writerow(["Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")])
    writer.writerow(["Best Concurrency", summary.get("best_concurrency", "N/A")])
    writer.writerow(["Total Requests", summary.get("total_requests", "N/A")])
    writer.writerow(["Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")])
    if summary.get("avg_goodput_percent") is not None:
        writer.writerow(["Avg Goodput (%)", summary.get("avg_goodput_percent")])
    writer.writerow([])

    # Results table
    writer.writerow(["# Detailed Results by Concurrency"])
    writer.writerow([
        "Concurrency",
        "Throughput (tok/s)",
        "Request Rate (req/s)",
        "TTFT p50 (ms)",
        "TTFT p95 (ms)",
        "TTFT p99 (ms)",
        "TPOT p50 (ms)",
        "TPOT p95 (ms)",
        "TPOT p99 (ms)",
        "E2E p50 (ms)",
        "E2E p95 (ms)",
        "E2E p99 (ms)",
        "Total Requests",
        "Successful",
        "Failed",
        "Error Rate (%)",
        "Goodput (%)",
    ])

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        writer.writerow([
            r.get("concurrency", "N/A"),
            f"{r.get('throughput_tokens_per_sec', 0):.2f}",
            f"{r.get('request_rate_per_sec', 0):.2f}",
            f"{ttft.get('p50', 0):.2f}",
            f"{ttft.get('p95', 0):.2f}",
            f"{ttft.get('p99', 0):.2f}",
            f"{tpot.get('p50', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p95', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p99', 0):.2f}" if tpot else "N/A",
            f"{e2e.get('p50', 0):.2f}",
            f"{e2e.get('p95', 0):.2f}",
            f"{e2e.get('p99', 0):.2f}",
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            f"{r.get('error_rate_percent', 0):.2f}",
            f"{goodput.get('goodput_percent', 0):.2f}" if goodput else "N/A",
        ])

    return output.getvalue()


def _export_to_xlsx(result: dict) -> bytes:
    """Convert benchmark result to Excel format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Install openpyxl: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    row = 1

    # Metadata section
    ws.cell(row=row, column=1, value="Benchmark Result Export").font = Font(bold=True, size=14)
    row += 2

    metadata = [
        ("Run ID", result.get("run_id", "N/A")),
        ("Model", result.get("model", "N/A")),
        ("Server URL", result.get("server_url", "N/A")),
        ("Adapter", result.get("adapter", "N/A")),
        ("Started At", result.get("started_at", "N/A")),
        ("Completed At", result.get("completed_at", "N/A")),
        ("Duration (s)", result.get("duration_seconds", "N/A")),
    ]

    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Summary section
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1

    summary = result.get("summary", {})
    summary_data = [
        ("Best Throughput (tok/s)", summary.get("best_throughput", "N/A")),
        ("Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")),
        ("Best Concurrency", summary.get("best_concurrency", "N/A")),
        ("Total Requests", summary.get("total_requests", "N/A")),
        ("Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")),
    ]
    if summary.get("avg_goodput_percent") is not None:
        summary_data.append(("Avg Goodput (%)", summary.get("avg_goodput_percent")))

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Results table
    ws.cell(row=row, column=1, value="Detailed Results by Concurrency").font = Font(bold=True, size=12)
    row += 1

    headers = [
        "Concurrency", "Throughput", "Req Rate", "TTFT p50", "TTFT p95", "TTFT p99",
        "TPOT p50", "TPOT p95", "TPOT p99", "E2E p50", "E2E p95", "E2E p99",
        "Total", "Success", "Failed", "Error %", "Goodput %",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row += 1

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        values = [
            r.get("concurrency", 0),
            round(r.get("throughput_tokens_per_sec", 0), 2),
            round(r.get("request_rate_per_sec", 0), 2),
            round(ttft.get("p50", 0), 2),
            round(ttft.get("p95", 0), 2),
            round(ttft.get("p99", 0), 2),
            round(tpot.get("p50", 0), 2) if tpot else None,
            round(tpot.get("p95", 0), 2) if tpot else None,
            round(tpot.get("p99", 0), 2) if tpot else None,
            round(e2e.get("p50", 0), 2),
            round(e2e.get("p95", 0), 2),
            round(e2e.get("p99", 0), 2),
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            round(r.get("error_rate_percent", 0), 2),
            round(goodput.get("goodput_percent", 0), 2) if goodput else None,
        ]

        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 12

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
