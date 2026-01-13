"""Benchmark API routes."""

import csv
import io
import json
from datetime import datetime
from typing import Literal, Optional

import httpx
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import Response, StreamingResponse

from llm_loadtest_api.auth import APIKeyAuth
from shared.database import Database
from llm_loadtest_api.logging_config import get_logger, log_benchmark_event
from llm_loadtest_api.services.benchmark_service import BenchmarkService
from llm_loadtest_api.models.schemas import (
    BenchmarkRequest,
    BenchmarkResponse,
    BenchmarkStatus,
    CompareRequest,
    CompareResponse,
    HealthResponse,
    RunListResponse,
)
from llm_loadtest_api import __version__

# GPU 모니터링
try:
    from shared.core.gpu_monitor import get_gpu_info
    GPU_AVAILABLE = True
except ImportError:
    GPU_AVAILABLE = False
    get_gpu_info = None

# Logger
logger = get_logger(__name__)


router = APIRouter(prefix="/api/v1/benchmark", tags=["benchmark"])

# Database and service instances (singleton pattern)
_db: Optional[Database] = None
_service: Optional[BenchmarkService] = None


def get_db() -> Database:
    """Get database instance."""
    global _db
    if _db is None:
        import os
        db_path = os.environ.get("DATABASE_PATH", "benchmarks.db")
        _db = Database(db_path)
    return _db


def get_service() -> BenchmarkService:
    """Get benchmark service instance."""
    global _service
    if _service is None:
        _service = BenchmarkService(get_db())
    return _service


@router.get("/health", response_model=HealthResponse)
async def health_check() -> HealthResponse:
    """Check API health."""
    return HealthResponse(
        status="healthy",
        version=__version__,
        timestamp=datetime.now(),
    )


@router.post("/run", response_model=dict, dependencies=[Depends(APIKeyAuth(required=True))])
async def start_benchmark(
    request: BenchmarkRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Start a new benchmark run.

    Requires API key authentication via X-API-Key header.

    Returns the run_id for tracking progress.
    """
    logger.info(
        "benchmark_request_received",
        server_url=request.server_url,
        model=request.model,
        adapter=request.adapter,
    )

    run_id = await service.start_benchmark(request)

    log_benchmark_event(
        "benchmark_started",
        run_id=run_id,
        extra={
            "server_url": request.server_url,
            "model": request.model,
            "adapter": request.adapter,
        },
    )

    return {"run_id": run_id, "status": "started"}


@router.get("/run/{run_id}", response_model=BenchmarkStatus)
async def get_run_status(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> BenchmarkStatus:
    """Get benchmark run status."""
    run = service.get_status(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")

    return BenchmarkStatus(
        run_id=run["id"],
        status=run["status"],
        server_url=run["server_url"],
        model=run["model"],
        adapter=run["adapter"],
        started_at=run.get("started_at"),
        completed_at=run.get("completed_at"),
        created_at=run["created_at"],
    )


@router.get("/result/{run_id}")
async def get_run_result(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Get benchmark result."""
    result = service.get_result(run_id)
    if not result:
        # Check if run exists but not completed
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    # Add summary if not present
    if "summary" not in result and "results" in result:
        results = result["results"]
        if results:
            best_throughput = max(r["throughput_tokens_per_sec"] for r in results)
            best_ttft = min(r["ttft"]["p50"] for r in results)
            best_result = max(results, key=lambda r: r["throughput_tokens_per_sec"])
            total_requests = sum(r["total_requests"] for r in results)
            failed_requests = sum(r["failed_requests"] for r in results)

            result["summary"] = {
                "best_throughput": best_throughput,
                "best_ttft_p50": best_ttft,
                "best_concurrency": best_result["concurrency"],
                "total_requests": total_requests,
                "overall_error_rate": (failed_requests / total_requests * 100) if total_requests > 0 else 0,
            }

            # Add Goodput if available
            goodput_results = [r["goodput"] for r in results if r.get("goodput")]
            if goodput_results:
                avg_goodput = sum(g["goodput_percent"] for g in goodput_results) / len(goodput_results)
                result["summary"]["avg_goodput_percent"] = avg_goodput

    return result


@router.get("/history", response_model=RunListResponse)
async def list_runs(
    limit: int = 50,
    offset: int = 0,
    status: Optional[str] = None,
    service: BenchmarkService = Depends(get_service),
) -> RunListResponse:
    """List benchmark runs."""
    runs = service.list_runs(limit, offset, status)

    return RunListResponse(
        runs=[
            BenchmarkStatus(
                run_id=r["id"],
                status=r["status"],
                server_url=r["server_url"],
                model=r["model"],
                adapter=r["adapter"],
                started_at=r.get("started_at"),
                completed_at=r.get("completed_at"),
                created_at=r["created_at"],
            )
            for r in runs
        ],
        total=len(runs),  # TODO: Get actual total count
        limit=limit,
        offset=offset,
    )


@router.delete("/run/{run_id}", dependencies=[Depends(APIKeyAuth(required=True))])
async def delete_run(
    run_id: str,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Delete a benchmark run.

    Requires API key authentication via X-API-Key header.
    """
    logger.info("benchmark_delete_request", run_id=run_id)

    deleted = service.delete_run(run_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Run not found")

    log_benchmark_event("benchmark_deleted", run_id=run_id)

    return {"deleted": run_id}


@router.post("/compare")
async def compare_runs(
    request: CompareRequest,
    service: BenchmarkService = Depends(get_service),
) -> dict:
    """Compare multiple benchmark runs."""
    comparison = service.compare_runs(request.run_ids)
    return comparison


@router.get("/result/{run_id}/export")
async def export_result(
    run_id: str,
    format: Literal["csv", "xlsx"] = "csv",
    service: BenchmarkService = Depends(get_service),
) -> Response:
    """Export benchmark result to CSV or Excel format.

    Args:
        run_id: The benchmark run ID.
        format: Export format ('csv' or 'xlsx').

    Returns:
        File download response.
    """
    result = service.get_result(run_id)
    if not result:
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    if format == "csv":
        content = _export_to_csv(result)
        return Response(
            content=content,
            media_type="text/csv",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.csv"'
            },
        )
    else:
        content = _export_to_xlsx(result)
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{run_id[:8]}_benchmark.xlsx"'
            },
        )


def _export_to_csv(result: dict) -> str:
    """Convert benchmark result to CSV format."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Metadata section
    writer.writerow(["# Benchmark Result Export"])
    writer.writerow(["Run ID", result.get("run_id", "N/A")])
    writer.writerow(["Model", result.get("model", "N/A")])
    writer.writerow(["Server URL", result.get("server_url", "N/A")])
    writer.writerow(["Adapter", result.get("adapter", "N/A")])
    writer.writerow(["Started At", result.get("started_at", "N/A")])
    writer.writerow(["Completed At", result.get("completed_at", "N/A")])
    writer.writerow(["Duration (s)", result.get("duration_seconds", "N/A")])
    writer.writerow([])

    # Summary section
    summary = result.get("summary", {})
    writer.writerow(["# Summary"])
    writer.writerow(["Best Throughput (tok/s)", summary.get("best_throughput", "N/A")])
    writer.writerow(["Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")])
    writer.writerow(["Best Concurrency", summary.get("best_concurrency", "N/A")])
    writer.writerow(["Total Requests", summary.get("total_requests", "N/A")])
    writer.writerow(["Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")])
    if summary.get("avg_goodput_percent") is not None:
        writer.writerow(["Avg Goodput (%)", summary.get("avg_goodput_percent")])
    writer.writerow([])

    # Results table
    writer.writerow(["# Detailed Results by Concurrency"])
    writer.writerow([
        "Concurrency",
        "Throughput (tok/s)",
        "Request Rate (req/s)",
        "TTFT p50 (ms)",
        "TTFT p95 (ms)",
        "TTFT p99 (ms)",
        "TPOT p50 (ms)",
        "TPOT p95 (ms)",
        "TPOT p99 (ms)",
        "E2E p50 (ms)",
        "E2E p95 (ms)",
        "E2E p99 (ms)",
        "Total Requests",
        "Successful",
        "Failed",
        "Error Rate (%)",
        "Goodput (%)",
    ])

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        writer.writerow([
            r.get("concurrency", "N/A"),
            f"{r.get('throughput_tokens_per_sec', 0):.2f}",
            f"{r.get('request_rate_per_sec', 0):.2f}",
            f"{ttft.get('p50', 0):.2f}",
            f"{ttft.get('p95', 0):.2f}",
            f"{ttft.get('p99', 0):.2f}",
            f"{tpot.get('p50', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p95', 0):.2f}" if tpot else "N/A",
            f"{tpot.get('p99', 0):.2f}" if tpot else "N/A",
            f"{e2e.get('p50', 0):.2f}",
            f"{e2e.get('p95', 0):.2f}",
            f"{e2e.get('p99', 0):.2f}",
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            f"{r.get('error_rate_percent', 0):.2f}",
            f"{goodput.get('goodput_percent', 0):.2f}" if goodput else "N/A",
        ])

    return output.getvalue()


def _export_to_xlsx(result: dict) -> bytes:
    """Convert benchmark result to Excel format."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export not available. Install openpyxl: pip install openpyxl",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmark Results"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    row = 1

    # Metadata section
    ws.cell(row=row, column=1, value="Benchmark Result Export").font = Font(bold=True, size=14)
    row += 2

    metadata = [
        ("Run ID", result.get("run_id", "N/A")),
        ("Model", result.get("model", "N/A")),
        ("Server URL", result.get("server_url", "N/A")),
        ("Adapter", result.get("adapter", "N/A")),
        ("Started At", result.get("started_at", "N/A")),
        ("Completed At", result.get("completed_at", "N/A")),
        ("Duration (s)", result.get("duration_seconds", "N/A")),
    ]

    for label, value in metadata:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    # Summary section
    ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=12)
    row += 1

    summary = result.get("summary", {})
    summary_data = [
        ("Best Throughput (tok/s)", summary.get("best_throughput", "N/A")),
        ("Best TTFT p50 (ms)", summary.get("best_ttft_p50", "N/A")),
        ("Best Concurrency", summary.get("best_concurrency", "N/A")),
        ("Total Requests", summary.get("total_requests", "N/A")),
        ("Overall Error Rate (%)", summary.get("overall_error_rate", "N/A")),
    ]
    if summary.get("avg_goodput_percent") is not None:
        summary_data.append(("Avg Goodput (%)", summary.get("avg_goodput_percent")))

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 2

    # Results table
    ws.cell(row=row, column=1, value="Detailed Results by Concurrency").font = Font(bold=True, size=12)
    row += 1

    headers = [
        "Concurrency", "Throughput", "Req Rate", "TTFT p50", "TTFT p95", "TTFT p99",
        "TPOT p50", "TPOT p95", "TPOT p99", "E2E p50", "E2E p95", "E2E p99",
        "Total", "Success", "Failed", "Error %", "Goodput %",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row += 1

    for r in result.get("results", []):
        ttft = r.get("ttft", {})
        tpot = r.get("tpot", {})
        e2e = r.get("e2e_latency", {})
        goodput = r.get("goodput", {})

        values = [
            r.get("concurrency", 0),
            round(r.get("throughput_tokens_per_sec", 0), 2),
            round(r.get("request_rate_per_sec", 0), 2),
            round(ttft.get("p50", 0), 2),
            round(ttft.get("p95", 0), 2),
            round(ttft.get("p99", 0), 2),
            round(tpot.get("p50", 0), 2) if tpot else None,
            round(tpot.get("p95", 0), 2) if tpot else None,
            round(tpot.get("p99", 0), 2) if tpot else None,
            round(e2e.get("p50", 0), 2),
            round(e2e.get("p95", 0), 2),
            round(e2e.get("p99", 0), 2),
            r.get("total_requests", 0),
            r.get("successful_requests", 0),
            r.get("failed_requests", 0),
            round(r.get("error_rate_percent", 0), 2),
            round(goodput.get("goodput_percent", 0), 2) if goodput else None,
        ]

        for col, value in enumerate(values, 1):
            ws.cell(row=row, column=col, value=value)
        row += 1

    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 12

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# Language instructions for AI analysis report
LANGUAGE_INSTRUCTIONS = {
    "ko": {
        "system": "한국어로 답변하세요.",
        "prompt_suffix": "",  # Korean is the default, no suffix needed
    },
    "en": {
        "system": "You MUST write your entire response in English. Do not use Korean.",
        "prompt_suffix": "\n\n---\n\n**[LANGUAGE REQUIREMENT - CRITICAL]**\n\nYou MUST write the entire analysis report in **English**.\n- All section headings must be in English\n- All explanations and analysis must be in English\n- All conclusions and recommendations must be in English\n- Do NOT use Korean anywhere in your response\n\nThis is a strict requirement. Respond ONLY in English.",
    },
    "zh": {
        "system": "你必须用中文撰写整个回复。不要使用韩语或英语。",
        "prompt_suffix": "\n\n---\n\n**[语言要求 - 重要]**\n\n你必须用**中文**撰写整个分析报告。\n- 所有章节标题必须使用中文\n- 所有解释和分析必须使用中文\n- 所有结论和建议必须使用中文\n- 不要在回复中使用韩语或英语\n\n这是严格的要求。请只用中文回复。",
    },
}


@router.get("/result/{run_id}/analysis")
async def analyze_result(
    run_id: str,
    server_url: str = Query(default="", description="vLLM server URL"),
    model: str = Query(default="", description="Model name (uses benchmark model if empty)"),
    is_thinking_model: bool = Query(default=False, description="Enable thinking model mode (buffers until </think> tag)"),
    language: str = Query(default="ko", description="Report language (ko, en, zh)"),
    service: BenchmarkService = Depends(get_service),
) -> StreamingResponse:
    """Generate AI analysis of benchmark results using vLLM.

    Streams the analysis response in real-time using Server-Sent Events (SSE).

    Args:
        run_id: The benchmark run ID.
        server_url: vLLM server URL for analysis generation.
        model: Model to use for analysis (defaults to benchmark's model).
        is_thinking_model: If True, waits for </think> tag before streaming. If False, streams immediately.

    Returns:
        StreamingResponse with SSE format.
    """
    result = service.get_result(run_id)
    if not result:
        run = service.get_status(run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        if run["status"] == "running":
            raise HTTPException(status_code=202, detail="Benchmark still running")
        if run["status"] == "failed":
            raise HTTPException(status_code=500, detail="Benchmark failed")
        raise HTTPException(status_code=404, detail="Result not found")

    # Check framework - AI analysis only supports vLLM
    framework = result.get("framework", "vllm")
    if framework and framework != "vllm":
        raise HTTPException(
            status_code=400,
            detail=f"AI analysis currently only supports vLLM framework. (Current: {framework})"
        )

    # Use model from result if not specified
    analysis_model = model if model else result.get("model", "qwen3-14b")

    # Build analysis prompt
    base_prompt = _build_analysis_prompt(result)

    # Get language settings
    lang_settings = LANGUAGE_INSTRUCTIONS.get(language, LANGUAGE_INSTRUCTIONS["ko"])
    language_system = lang_settings["system"]
    prompt_suffix = lang_settings["prompt_suffix"]

    # Add language suffix to prompt if needed
    prompt = base_prompt + prompt_suffix

    async def generate_analysis():
        """Stream analysis from vLLM."""
        # System prompt with language instruction
        base_system_prompt = f"""당신은 LLM 서버 성능 분석 전문가입니다. 벤치마크 결과를 분석하여 마크다운 형식의 보고서를 작성합니다.

        [답변 원칙]
        - {language_system}
        - 전문 용어는 괄호 안에 간단한 설명 추가 (예: TTFT(첫 토큰 응답 시간))
        - 구조화된 마크다운 형식 사용
        - 핵심부터 설명, 불필요한 서론 생략"""

        # Thinking 모델이 아니면 /no_think 지시어 추가
        if is_thinking_model:
            system_prompt = base_system_prompt
        else:
            system_prompt = "/no_think\n" + base_system_prompt

        try:
            # Thinking 모델일 때만 thinking 상태 전송
            if is_thinking_model:
                yield f"data: {json.dumps({'thinking': True})}\n\n"

            async with httpx.AsyncClient(timeout=httpx.Timeout(300.0, read=300.0)) as client:
                async with client.stream(
                    "POST",
                    f"{server_url}/v1/chat/completions",
                    json={
                        "model": analysis_model,
                        "messages": [
                            {
                                "role": "system",
                                "content": system_prompt
                            },
                            {
                                "role": "user",
                                "content": prompt
                            }
                        ],
                        "stream": True,
                        "max_tokens": 8192,
                        "temperature": 0.3,
                    },
                    headers={"Content-Type": "application/json"},
                ) as response:
                    if response.status_code != 200:
                        error_text = await response.aread()
                        yield f"data: {json.dumps({'error': f'vLLM error: {error_text.decode()}'})}\n\n"
                        return

                    if is_thinking_model:
                        # Thinking 모델: </think> 태그까지 버퍼링
                        buffer = ""
                        report_started = False
                        think_end_tag = "</think>"

                        async for line in response.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]":
                                    # 버퍼에 남은 내용이 있으면 출력
                                    if buffer and not report_started:
                                        yield f"data: {json.dumps({'thinking': False})}\n\n"
                                        yield f"data: {json.dumps({'content': buffer})}\n\n"
                                    yield "data: [DONE]\n\n"
                                    break
                                try:
                                    chunk = json.loads(data)
                                    content = chunk.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                    if content:
                                        if report_started:
                                            yield f"data: {json.dumps({'content': content})}\n\n"
                                        else:
                                            buffer += content
                                            # </think> 태그가 나오면 보고서 시작
                                            if think_end_tag in buffer:
                                                idx = buffer.find(think_end_tag)
                                                remaining = buffer[idx + len(think_end_tag):].lstrip()
                                                report_started = True
                                                yield f"data: {json.dumps({'thinking': False})}\n\n"
                                                if remaining:
                                                    yield f"data: {json.dumps({'content': remaining})}\n\n"
                                                buffer = ""
                                except json.JSONDecodeError:
                                    continue
                    else:
                        # Non-thinking 모델: 바로 스트리밍
                        async for line in response.aiter_lines():
                            if line.startswith("data: "):
                                data = line[6:]
                                if data == "[DONE]":
                                    yield "data: [DONE]\n\n"
                                    break
                                try:
                                    chunk = json.loads(data)
                                    content = chunk.get("choices", [{}])[0].get("delta", {}).get("content", "")
                                    if content:
                                        yield f"data: {json.dumps({'content': content})}\n\n"
                                except json.JSONDecodeError:
                                    continue
        except httpx.ConnectError:
            yield f"data: {json.dumps({'error': f'Cannot connect to vLLM server: {server_url}'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        generate_analysis(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def _build_analysis_prompt(result: dict) -> str:
    """Build analysis prompt from benchmark result with infrastructure context."""
    model = result.get("model", "Unknown")
    server_url = result.get("server_url", "Unknown")
    duration = result.get("duration_seconds", 0)
    results = result.get("results", [])
    config = result.get("config", {})

    # Get server infrastructure from saved result
    server_infra = result.get("server_infra")

    # Build infrastructure section
    infra_section = _build_infra_section(server_infra)

    # Build workload section from config
    workload_section = _build_workload_section(config)

    # Calculate summary statistics from results
    best_throughput = 0.0
    best_ttft_p50 = float('inf')
    best_concurrency = None
    total_errors = 0
    total_requests = 0
    goodput_values = []

    for r in results:
        throughput = r.get('throughput_tokens_per_sec', 0) or 0
        if throughput > best_throughput:
            best_throughput = throughput
            best_concurrency = r.get('concurrency', 0)

        ttft = r.get("ttft") or {}
        ttft_p50 = ttft.get('p50', 0) or 0
        if ttft_p50 > 0 and ttft_p50 < best_ttft_p50:
            best_ttft_p50 = ttft_p50

        total_requests += r.get('total_requests', 0) or 0
        total_errors += r.get('error_count', 0) or 0

        goodput = r.get("goodput") or {}
        if goodput and 'goodput_percent' in goodput:
            goodput_values.append(goodput['goodput_percent'])

    overall_error_rate = (total_errors / total_requests * 100) if total_requests > 0 else 0
    avg_goodput = sum(goodput_values) / len(goodput_values) if goodput_values else None
    if best_ttft_p50 == float('inf'):
        best_ttft_p50 = 0

    # Build concurrency results table
    results_table = _build_results_table(results)

    goodput_summary = f"{avg_goodput:.1f}%" if avg_goodput is not None else "N/A"
    concurrency_summary = best_concurrency if best_concurrency else "N/A"

    # Build GPU metrics note based on whether vllm_config was provided
    vllm_config = server_infra.get("vllm_config", {}) if server_infra else {}
    gpu_mem_util = vllm_config.get("gpu_memory_utilization") if vllm_config else None

    if gpu_mem_util:
        gpu_metrics_note = f"""
        **참고**: 사용자가 제공한 vLLM 설정에 따르면 `gpu_memory_utilization={gpu_mem_util}` 입니다.
        GPU 메모리의 {gpu_mem_util * 100:.0f}%만 vLLM에 할당되므로, GPU 메트릭 해석 시 이를 고려하세요.
        예: GPU 메모리 사용률 {gpu_mem_util * 100 * 0.9:.0f}%는 vLLM 할당분의 약 90%를 사용 중입니다.
        """
    else:
        gpu_metrics_note = """
        **주의**: GPU 메트릭은 시스템 전체 사용량입니다. vLLM 설정값이 제공되지 않았으므로 기본값(`gpu_memory_utilization=0.9`)을 가정합니다.
        GPU 메모리의 90%만 vLLM에 할당됩니다. 예를 들어 GPU 메모리 사용률 85%는 vLLM 할당분의 거의 전부를 사용 중일 수 있습니다.
        """

    prompt = f"""
    다음 LLM 서버 벤치마크 결과를 분석해주세요.

    {infra_section}

    {workload_section}

    ## 테스트 요약
    - **모델**: {model}
    - **서버**: {server_url}
    - **테스트 시간**: {duration:.1f}초

    ## 성능 요약
    - **최고 처리량**: {best_throughput:.1f} tok/s (동시성 {concurrency_summary}에서)
    - **최저 TTFT p50**: {best_ttft_p50:.1f} ms
    - **전체 에러율**: {overall_error_rate:.2f}%
    - **평균 Goodput**: {goodput_summary}

    ## Concurrency별 상세 결과
    {results_table}

    ---

    ## 분석 요청

    위 벤치마크 결과와 **서버 인프라 환경을 고려하여** 전문가 보고서를 작성해주세요.

    **[용어 설명 규칙]**
    - 전문 용어는 처음 사용 시 괄호 안에 간단한 설명 추가
    - 예: "TTFT(첫 토큰 응답 시간)", "Throughput(처리량)", "Goodput(SLA 충족 비율)"

    **[분석 항목]**

    # 1. 환경 요약
    서버 인프라(GPU, CPU, 메모리)와 테스트 환경을 간략히 정리하세요.

    # 2. 성능 개요
    GPU 스펙 대비 성능을 평가하세요. 처리량과 응답 시간의 전반적인 수준을 분석하세요.

    # 3. Concurrency 영향 분석
    동시성 증가에 따른 성능 변화 패턴을 분석하세요. 표 데이터를 기반으로 구체적인 수치를 인용하세요.

    # 4. 병목 원인 분석 (증거 기반)
    성능 저하 지점을 식별하고, 데이터를 근거로 병목 원인을 분석하세요.

    {gpu_metrics_note}

    다음 형식을 따르세요:
    - **가설**: [성능 저하 원인 추정]
    - **증거**: [GPU 메트릭, 에러율, TTFT/처리량 변화 등 데이터 인용]
    - **검증**: [다른 가능한 원인 배제 또는 추가 확인 필요 사항]

    # 5. 권장 운영 동시성
    최적 동시성 레벨과 그 근거를 제시하세요. SLA(예: TTFT p99 < 500ms) 충족 여부를 고려하세요.

    # 6. 개선 제안
    다음 측면에서 우선순위가 높은 개선안을 제시하세요:
    - **인프라**: GPU 추가, 메모리 최적화
    - **설정**: max_num_seqs, gpu_memory_utilization 조정
    - **모델**: 양자화, context length 최적화

    # 7. 모니터링 권장안
    운영 시 모니터링해야 할 주요 메트릭과 알람 임계값을 제안하세요.

    각 항목을 **마크다운 헤딩(#)으로 구분**하여 작성해주세요.
    """

    return prompt


def _build_infra_section(server_infra: dict | None) -> str:
    """Build infrastructure section for prompt."""
    if not server_infra:
        # Fallback: try to get current GPU info (legacy behavior)
        if GPU_AVAILABLE and get_gpu_info:
            try:
                gpu_result = get_gpu_info()
                if gpu_result.available and gpu_result.metrics:
                    gpu_lines = []
                    for gpu in gpu_result.metrics:
                        gpu_lines.append(
                            f"  - **GPU {gpu.gpu_index}**: {gpu.device_name} "
                            f"({gpu.memory_total_gb:.1f}GB VRAM)"
                        )
                    return f"""## 1. 서버 인프라
> (현재 시점 GPU 정보 - 벤치마크 시점과 다를 수 있음)

### 1.1 GPU
- **GPU 수**: {gpu_result.gpu_count}장
{chr(10).join(gpu_lines)}"""
            except Exception:
                pass
        return "## 1. 서버 인프라\n(인프라 정보 없음 - 일반적인 분석만 가능)"

    lines = ["## 1. 서버 인프라"]

    # GPU info
    gpu = server_infra.get("gpu", {})
    if gpu:
        lines.append("\n### 1.1 GPU")
        lines.append(f"- **모델**: {gpu.get('gpu_model', 'Unknown')}")
        lines.append(f"- **개수**: {gpu.get('gpu_count', 'Unknown')}장")
        lines.append(f"- **VRAM**: {gpu.get('gpu_memory_total_gb', 'Unknown')} GB/장")
        if gpu.get("driver_version"):
            lines.append(f"- **드라이버**: {gpu['driver_version']}")
        if gpu.get("cuda_version"):
            lines.append(f"- **CUDA**: {gpu['cuda_version']}")
        if gpu.get("mig_enabled") is not None:
            lines.append(f"- **MIG**: {'활성화' if gpu['mig_enabled'] else '비활성화'}")

    # System info
    system = server_infra.get("system", {})
    if system:
        lines.append("\n### 1.2 System")
        lines.append(f"- **CPU**: {system.get('cpu_model', 'Unknown')}")
        cores_physical = system.get('cpu_cores_physical', '?')
        cores_logical = system.get('cpu_cores_logical', '?')
        lines.append(f"- **코어**: {cores_physical} physical / {cores_logical} logical")
        lines.append(f"- **RAM**: {system.get('ram_total_gb', 'Unknown')} GB")

    # Serving engine
    engine = server_infra.get("serving_engine", {})
    if engine:
        lines.append("\n### 1.3 Serving Engine")
        lines.append(f"- **엔진**: {engine.get('engine_type', 'Unknown')}")
        if engine.get("engine_version"):
            lines.append(f"- **버전**: {engine['engine_version']}")
        if engine.get("model_name"):
            lines.append(f"- **모델**: {engine['model_name']}")
        if engine.get("quantization"):
            lines.append(f"- **양자화**: {engine['quantization']}")
        if engine.get("max_num_seqs"):
            lines.append(f"- **max_num_seqs**: {engine['max_num_seqs']}")
        if engine.get("gpu_memory_utilization"):
            lines.append(f"- **gpu_memory_utilization**: {engine['gpu_memory_utilization']}")
        if engine.get("tensor_parallel_size"):
            lines.append(f"- **tensor_parallel_size**: {engine['tensor_parallel_size']}")

    # User-provided vLLM configuration
    vllm_config = server_infra.get("vllm_config", {})
    if vllm_config:
        lines.append("\n### 1.4 vLLM 설정 (사용자 제공)")
        if vllm_config.get("gpu_memory_utilization") is not None:
            lines.append(f"- **gpu_memory_utilization**: {vllm_config['gpu_memory_utilization']}")
        if vllm_config.get("tensor_parallel_size") is not None:
            lines.append(f"- **tensor_parallel_size**: {vllm_config['tensor_parallel_size']}")
        if vllm_config.get("max_num_seqs") is not None:
            lines.append(f"- **max_num_seqs**: {vllm_config['max_num_seqs']}")
        if vllm_config.get("quantization") is not None:
            lines.append(f"- **quantization**: {vllm_config['quantization']}")

    # GPU metrics during benchmark
    gpu_metrics = server_infra.get("gpu_metrics_during_benchmark", [])
    if gpu_metrics:
        section_num = "1.5" if vllm_config else "1.4"
        lines.append(f"\n### {section_num} 벤치마크 중 GPU 상태")

        # Add note about GPU metrics interpretation
        gpu_mem_util = vllm_config.get("gpu_memory_utilization") if vllm_config else None
        if gpu_mem_util:
            lines.append(f"> **참고**: vLLM이 GPU 메모리의 {gpu_mem_util * 100:.0f}%만 사용하도록 설정됨")
            lines.append("> 아래 메트릭은 시스템 전체 사용량이며, vLLM 할당분 대비 계산이 필요합니다.")
        else:
            lines.append("> **주의**: 아래 GPU 메트릭은 시스템 전체 사용량입니다.")
            lines.append("> vLLM의 `gpu_memory_utilization` 설정(기본값 0.9)으로 실제 할당 가능 메모리가 제한될 수 있습니다.")
        lines.append("")

        for m in gpu_metrics:
            avg_mem = m.get('avg_memory_used_gb', 0)
            peak_mem = m.get('peak_memory_used_gb', 0)
            avg_util = m.get('avg_gpu_util_percent', 0)
            peak_util = m.get('peak_gpu_util_percent', 0)
            lines.append(
                f"- **GPU {m.get('gpu_index', 0)}**: "
                f"메모리 {avg_mem:.1f}GB (peak {peak_mem:.1f}GB), "
                f"활용률 {avg_util:.0f}% (peak {peak_util:.0f}%)"
            )

    return "\n".join(lines)


def _build_workload_section(config: dict) -> str:
    """Build workload section for prompt."""
    if not config:
        return ""

    lines = ["## 2. 실험 설계"]
    lines.append("\n### 2.1 Workload")
    lines.append(f"- **Input tokens**: {config.get('input_len', 'Unknown')}")
    lines.append(f"- **Output tokens**: {config.get('output_len', 'Unknown')}")
    lines.append(f"- **요청 수**: {config.get('num_prompts', 'Unknown')}/레벨")
    concurrency = config.get('concurrency', [])
    lines.append(f"- **동시성 레벨**: {concurrency}")
    lines.append(f"- **스트리밍**: {'예' if config.get('stream', True) else '아니오'}")

    # SLA if defined
    thresholds = config.get("goodput_thresholds", {})
    if thresholds:
        lines.append("\n### 2.2 SLA 정의")
        if thresholds.get("ttft_ms"):
            lines.append(f"- **TTFT**: ≤ {thresholds['ttft_ms']}ms")
        if thresholds.get("tpot_ms"):
            lines.append(f"- **TPOT**: ≤ {thresholds['tpot_ms']}ms")
        if thresholds.get("e2e_ms"):
            lines.append(f"- **E2E**: ≤ {thresholds['e2e_ms']}ms")

    return "\n".join(lines)


def _build_results_table(results: list) -> str:
    """Build markdown table of concurrency results."""
    table = "| Concurrency | Throughput (tok/s) | TTFT p50 (ms) | TTFT p99 (ms) | Error Rate (%) | Goodput (%) |\n"
    table += "|-------------|-------------------|---------------|---------------|----------------|-------------|\n"

    for r in results:
        ttft = r.get("ttft") or {}
        goodput = r.get("goodput") or {}
        goodput_str = f"{goodput.get('goodput_percent', 0):.1f}" if goodput else "N/A"
        table += (
            f"| {r.get('concurrency', 0)} "
            f"| {r.get('throughput_tokens_per_sec', 0):.1f} "
            f"| {ttft.get('p50', 0):.1f} "
            f"| {ttft.get('p99', 0):.1f} "
            f"| {r.get('error_rate_percent', 0):.2f} "
            f"| {goodput_str} |\n"
        )

    return table
